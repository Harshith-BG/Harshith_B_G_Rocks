import openpyxl as xl
import random

wb = xl.load_workbook('BOOK.xlsx')
sheet = wb['Sheet1']


def main():
    mcf = MiniChess()

    mcf.before_play()
    mcf.choose_turn()
    mcf.main_content()

    mcf.win_or_loss()
    mcf.no_duplicates()


class MiniChess:
    def __init__(self):
        self.board = ['a', 'b', 'c', ' ', ' ', ' ', 'x', 'y', 'z']
        self.n = 0
        self.list1 = ['a', 'b', 'c']
        self.list2 = ['x', 'y', 'z']
        self.record = [-1, -1, -1, -1]
        self.r0w = 2
        self.played_moves = []
        self.played_moves2 = []
        self.recorded_moves = []
        self.best_moves = []
        self.best_moves2 = []
        self.condition = None
        self.iteration_count = None
        self.element = None
        self.invert_or = 1

        self.list_used = None
        self.times = None
        self.output = None
        self.position = None
        self.position1 = None
        self.position2 = None
        self.action = None
        self.choice = None
        self.choice2 = None
        self.times1 = None
        self.times2 = None
        self.storage = None
        self.col = 1
        self.counter = 0
        self.row2 = None
        self.col2 = None

    def before_play(self):
        while sheet.cell(self.r0w, column=8).value in [1, -1, 0]:
            self.r0w += 1

    def draw_board(self):
        self.times = 0
        self.output = '         | '
        print()
        print("         =============")
        for block in self.board:
            self.output += block + ' | '
            self.times += 1
            if self.times % 3 == 0:
                print(self.output)
                print("         =============")
                self.output = '         | '

    def previous_moves(self):
        for self.col2 in range(1, 8):
            if sheet.cell(self.row2, self.col2).value in [1, 2, 3, 4]:
                self.recorded_moves.append(sheet.cell(self.row2, self.col2).value)
            else:
                break

    def take_input(self):
        # print(f'{self.times1},{self.times2}')
        if (self.n == 0 and self.invert_or == 1) or (self.n == 1 and self.invert_or == -1):
            print(f"Computer's choice :{self.choice}")
            self.move_accord_input()
        else:
            self.choice = input("enter your choice :")
            while self.choice.isalpha():
                print("ENTERED CHOICE IS INCORRECT")
                self.choice = input("enter your choice :")
            self.choice = int(self.choice)
            if self.choice < 1 or self.choice > self.times:
                print("ENTERED CHOICE IS INCORRECT")
                self.take_input()
            else:
                self.move_accord_input()
        self.played_moves.append(self.choice)
        cell = sheet.cell(self.r0w, self.col)
        cell.value = self.choice
        self.col += 1

    def move_accord_input(self):
        self.choice2 = self.record[self.choice - 1]
        if self.choice in range(self.times1 + 1):
            self.position = self.position1
        elif self.choice in range(self.times1 + 1, self.times1 + self.times2 + 1):
            self.position = self.position2

        if self.choice2 == 0:
            self.board[self.position + 3] = self.board[self.position]
        elif self.choice2 == 1:
            self.board[self.position - 3] = self.board[self.position]
        elif self.choice2 == 2:
            self.board[self.position + 2] = self.board[self.position]
        elif self.choice2 == 3:
            self.board[self.position - 4] = self.board[self.position]
        elif self.choice2 == 4:
            self.board[self.position + 4] = self.board[self.position]
        elif self.choice2 == 5:
            self.board[self.position - 2] = self.board[self.position]
        self.board[self.position] = ' '

    def possible_moves(self):
        self.times = 0
        self.times1 = 0
        self.times2 = 0
        if self.n == 0:
            self.list_used = self.list1
        else:
            self.list_used = self.list2

        for ele in self.list_used:
            for ele2 in self.board:
                if ele == ele2:
                    if ele2 == 'a' or ele2 == 'x':
                        self.position1 = self.board.index(ele)
                    elif ele2 == 'b' or ele2 == 'y':
                        self.position2 = self.board.index(ele)
                    self.storage = self.times
                    self.position = self.board.index(ele)

                    self.moves_forward()
                    if self.position % 3 == 1:
                        self.captures_left()
                        self.captures_right()
                    elif self.position % 3 == 0:
                        self.captures_right()
                    else:
                        self.captures_left()

                    if ele2 == 'a' or ele2 == 'x':
                        self.times1 = self.times - self.storage
                    elif ele2 == 'b' or ele2 == 'y':
                        self.times2 = self.times - self.storage

    def moves_forward(self):
        if self.n == 0:
            if self.board[self.position + 3] == ' ':
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} could move forward')
                self.record[self.times - 1] = 0
        else:
            if self.board[self.position - 3] == ' ':
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} could move forward')
                self.record[self.times - 1] = 1

    def captures_left(self):
        if self.n == 0:
            if self.board[self.position + 2] in self.list2:
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} captures {self.board[self.position + 2]}')
                self.record[self.times - 1] = 2
        else:
            if self.board[self.position - 4] in self.list1:
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} captures {self.board[self.position - 4]}')
                self.record[self.times - 1] = 3

    def captures_right(self):
        if self.n == 0:
            if self.board[self.position + 4] in self.list2:
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} captures {self.board[self.position + 4]}')
                self.record[self.times - 1] = 4
        else:
            if self.board[self.position - 2] in self.list1:
                self.times += 1
                print(f'ACTION[{self.times}] : {self.board[self.position]} captures {self.board[self.position - 2]}')
                self.record[self.times - 1] = 5

    def perfect_play(self):
        self.condition = 1
        for self.row2 in range(2, sheet.max_row + 1):
            self.previous_moves()
            if self.recorded_moves[:-1] == self.played_moves[:]:
                if sheet.cell(self.row2, 8).value == (self.condition * self.invert_or):
                    self.best_moves.append(sheet.cell(self.row2, len(self.played_moves) + 1).value)
                    break
            self.recorded_moves.clear()
            if self.row2 == sheet.max_row and self.condition == 1 and len(self.best_moves) < 1:
                self.condition = 0
                self.row2 = 2

        if self.condition == 0:
            for self.iteration_count in range(1, self.times + 1):
                self.best_moves.append(self.iteration_count)

            for self.row2 in range(2, sheet.max_row + 1):
                self.previous_moves()
                if self.recorded_moves[:-2] == self.played_moves[:]:
                    if sheet.cell(self.row2, 8).value == (-1 * self.invert_or):
                        self.best_moves.remove(sheet.cell(self.row2, len(self.played_moves) + 1).value)
                self.recorded_moves.clear()

        if ((len(self.played_moves) == 2 and self.invert_or == 1) or (
                len(self.played_moves) == 3 and self.invert_or == -1)) and len(self.best_moves) == 0:
            for element in self.best_moves:
                self.best_moves2.append(element)
            for element in self.played_moves:
                self.played_moves2.append(element)
            for element in self.best_moves2:
                self.played_moves2.append(element)
                self.iteration_count = 0
                for self.row2 in range(2, sheet.max_row + 1):
                    self.previous_moves()
                    if self.recorded_moves[:-3] == self.played_moves2[:]:
                        if sheet.cell(self.row2, 8).value == (-1 * self.invert_or):
                            # print(f'{sheet.cell(mcf.row2, 3).value},{best_moves}')
                            self.iteration_count += 1
                            # best_moves.remove(sheet.cell(mcf.row2, 3).value)
                    self.recorded_moves.clear()
                    if self.iteration_count > 1 and self.invert_or == 1:
                        self.best_moves.remove(sheet.cell(self.row2, 3).value)
                        break
                    elif self.iteration_count == 0 and self.invert_or == -1:
                        self.best_moves.clear()
                        self.best_moves.append(sheet.cell(self.row2, 3).value)
                        break
                self.played_moves2.pop()
            self.best_moves2.clear()
            self.played_moves2.clear()

    def no_duplicates(self):
        for self.row2 in range(2, sheet.max_row):
            self.previous_moves()
            if self.recorded_moves[:] == self.played_moves[:] and (
                    (not sheet.cell(sheet.max_row, 8).value == (-1 * self.invert_or)) or (
                    sheet.cell(sheet.max_row, 8) == 0)):
                for self.col2 in range(1, 9):
                    sheet.cell(sheet.max_row, self.col2).value = ''
                break
            else:
                self.recorded_moves.clear()

    def complete_delete(self):
        for self.row2 in range(2, sheet.max_row):
            for self.col2 in range(1, 9):
                sheet.cell(self.row2, self.col2).value = ''

    def win_or_loss(self):
        if self.board[0] in self.list2 or self.board[1] in self.list2 or self.board[2] in self.list2:
            print('''PLAYER WITH TEAM [X , Y , Z]
         HAS WON THE MATCH''')
            sheet.cell(self.r0w, 8).value = -1
        elif self.board[6] in self.list1 or self.board[7] in self.list1 or self.board[8] in self.list1:
            print('''PLAYER WITH TEAM [A , B , C]
         HAS WON THE MATCH''')
            sheet.cell(self.r0w, 8).value = 1
        else:
            self.times = 0
            for self.ele2 in self.board:
                if self.ele2 in self.list1:
                    self.times = 1
                    break
            if self.times == 0:
                print('''PLAYER WITH TEAM [X , Y , Z]
         HAS WON THE MATCH''')
                sheet.cell(self.r0w, 8).value = -1
            else:
                self.times = 0
                for self.ele2 in self.board:
                    if self.ele2 in self.list2:
                        self.times += 1
                        break
                if self.times == 0:
                    print('''PLAYER WITH TEAM [A , B , C]
         HAS WON THE MATCH''')
                    sheet.cell(self.r0w, 8).value = 1
                else:
                    print("   THIS MATCH WAS BORING")
                    sheet.cell(self.r0w, 8).value = 0

    def choose_turn(self):
        while True:
            self.invert_or = input('''ENTER 1 to play first 
ENTER 2 to play second
ENTER 3 to delete previous records
enter your choice :''')
            if self.invert_or.isdigit():
                self.invert_or = int(self.invert_or)
                if self.invert_or == 1:
                    self.invert_or = -1
                    self.draw_board()
                    break
                elif self.invert_or == 2:
                    self.invert_or = 1
                    self.draw_board()
                    break
                elif self.invert_or == 3:
                    self.complete_delete()
                    self.r0w = 2
                else:
                    print("WARNING: YOUR NUMBER SHOULD BE EITHER 1 or 2")
            else:
                print("""YOUR INPUT CAN'T BE AN ALPHABET 
    TRY ENTERING 1 OR 2""")
            print("")
        print("")

    def main_content(self):
        while (not (self.board[0] in self.list2 or self.board[1] in self.list2 or self.board[2] in self.list2)) and (
                not (
                        self.board[6] in self.list1 or self.board[7] in self.list1 or self.board[8] in self.list1)):
            self.possible_moves()
            if self.times == 0:
                break

            if (self.n == 0 and self.invert_or == 1) or (self.n == 1 and self.invert_or == -1):
                self.perfect_play()
                print(self.best_moves)
                if len(self.best_moves) == 0:
                    self.choice = random.randint(1, self.times)
                else:
                    self.choice = self.best_moves[random.randint(1, len(self.best_moves)) - 1]
                self.best_moves.clear()

            self.take_input()

            self.counter += 1
            print("\n           TURN : " + str(self.counter))

            self.draw_board()
            self.record.clear()
            self.record = [-1, -1, -1, -1]
            if self.n == 0:
                self.n += 1
            else:
                self.n -= 1


if __name__ == "__main__":
    main()
    wb.save('BOOK.xlsx')
